from io import BytesIO, StringIO
import csv
import os
from pathlib import Path
import re
import secrets
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook


MAX_FILES_PER_MESSAGE = 3
MAX_FILE_BYTES = 8 * 1024 * 1024
MAX_MESSAGE_FILE_BYTES = 15 * 1024 * 1024
MAX_CONVERSATION_BYTES = 50 * 1024 * 1024
ALLOWED_TYPES = {
    "application/pdf": ".pdf",
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/webp": ".webp",
    "text/csv": ".csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
}


def upload_root() -> Path:
    path = Path(os.getenv("AI_UPLOAD_DIR", "/data/ai-uploads")).resolve()
    path.mkdir(parents=True, exist_ok=True)
    return path


def _safe_original_name(value: str | None) -> str:
    name = Path(value or "attachment").name
    name = re.sub(r"[^A-Za-z0-9._()\- ก-๙]", "_", name).strip(". ")
    return (name or "attachment")[:255]


def _validate_signature(data: bytes, mime_type: str) -> None:
    valid = True
    if mime_type == "application/pdf":
        valid = data.startswith(b"%PDF-")
    elif mime_type == "image/png":
        valid = data.startswith(b"\x89PNG\r\n\x1a\n")
    elif mime_type == "image/jpeg":
        valid = data.startswith(b"\xff\xd8\xff")
    elif mime_type == "image/webp":
        valid = len(data) >= 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP"
    elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        valid = data.startswith(b"PK\x03\x04")
    if not valid:
        raise HTTPException(status_code=422, detail="Attachment content does not match its declared file type")


async def read_uploads(files: list[UploadFile]) -> list[dict[str, Any]]:
    if len(files) > MAX_FILES_PER_MESSAGE:
        raise HTTPException(status_code=413, detail=f"Attach at most {MAX_FILES_PER_MESSAGE} files per message")
    loaded: list[dict[str, Any]] = []
    total = 0
    for item in files:
        mime_type = (item.content_type or "").lower()
        if mime_type not in ALLOWED_TYPES:
            raise HTTPException(status_code=415, detail=f"Unsupported attachment type: {mime_type or 'unknown'}")
        data = await item.read(MAX_FILE_BYTES + 1)
        if not data:
            raise HTTPException(status_code=422, detail="Attachments cannot be empty")
        if len(data) > MAX_FILE_BYTES:
            raise HTTPException(status_code=413, detail="Each attachment must be 8 MB or smaller")
        total += len(data)
        if total > MAX_MESSAGE_FILE_BYTES:
            raise HTTPException(status_code=413, detail="Attachments must total 15 MB or less per message")
        _validate_signature(data, mime_type)
        loaded.append({
            "name": _safe_original_name(item.filename),
            "mime_type": mime_type,
            "data": data,
            "size_bytes": len(data),
            "suffix": ALLOWED_TYPES[mime_type],
        })
    return loaded


def save_upload(conversation_id: str, item: dict[str, Any]) -> tuple[str, str]:
    directory = upload_root() / conversation_id
    directory.mkdir(parents=True, exist_ok=True)
    stored_name = f"{secrets.token_hex(16)}{item['suffix']}"
    target = (directory / stored_name).resolve()
    if directory.resolve() not in target.parents:
        raise HTTPException(status_code=400, detail="Invalid attachment path")
    target.write_bytes(item["data"])
    return stored_name, str(target)


def remove_file(conversation_id: str, stored_name: str) -> None:
    directory = (upload_root() / conversation_id).resolve()
    target = (directory / Path(stored_name).name).resolve()
    if directory in target.parents and target.exists():
        target.unlink()


def remove_conversation_files(conversation_id: str) -> None:
    directory = (upload_root() / conversation_id).resolve()
    root = upload_root()
    if root not in directory.parents or not directory.exists():
        return
    for file in directory.iterdir():
        if file.is_file():
            file.unlink()
    directory.rmdir()


def attachment_for_gemini(conversation_id: str, stored_name: str, mime_type: str, original_name: str) -> dict[str, Any]:
    path = (upload_root() / conversation_id / Path(stored_name).name).resolve()
    directory = (upload_root() / conversation_id).resolve()
    if directory not in path.parents or not path.exists():
        raise HTTPException(status_code=410, detail=f"Attachment is no longer available: {original_name}")
    data = path.read_bytes()
    if mime_type == "text/csv":
        return {"kind": "text", "name": original_name, "text": _csv_text(data), "mime_type": mime_type}
    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return {"kind": "text", "name": original_name, "text": _xlsx_text(data), "mime_type": mime_type}
    return {"kind": "binary", "name": original_name, "data": data, "mime_type": mime_type}


def upload_for_gemini(item: dict[str, Any]) -> dict[str, Any]:
    mime_type = item["mime_type"]
    if mime_type == "text/csv":
        return {"kind": "text", "name": item["name"], "text": _csv_text(item["data"]), "mime_type": mime_type}
    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return {"kind": "text", "name": item["name"], "text": _xlsx_text(item["data"]), "mime_type": mime_type}
    return {"kind": "binary", "name": item["name"], "data": item["data"], "mime_type": mime_type}


def _csv_text(data: bytes) -> str:
    decoded = data.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(StringIO(decoded)))[:500]
    return "\n".join(" | ".join(cell[:500] for cell in row[:30]) for row in rows)


def _xlsx_text(data: bytes) -> str:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets[:8]:
        lines.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(max_row=300, max_col=30, values_only=True):
            lines.append(" | ".join("" if value is None else str(value)[:500] for value in row))
    workbook.close()
    return "\n".join(lines)[:80_000]
